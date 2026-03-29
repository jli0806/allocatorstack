#!/usr/bin/env python3
"""Generate Excel comparison matrices from JSON data.

Usage:
    python scripts/generate-excel.py workspace/ddq-output.json --output workspace/comparison.xlsx
    python scripts/generate-excel.py workspace/screening-output.json --output workspace/screening.xlsx

Output: Formatted .xlsx file with comparison matrix.
"""
import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Severity color coding
COLORS = {
    "CRITICAL": "FF4444",
    "WARNING": "FFAA00",
    "INFO": "4488FF",
    "PASS": "44AA44",
    "FAIL": "FF4444",
    "BORDERLINE": "FFAA00",
}


def build_comparison_matrix(data, wb: Workbook):
    """Build a comparison matrix sheet from structured JSON data."""
    ws = wb.active
    ws.title = "Comparison Matrix"

    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")

    # Handle both single-manager and multi-manager input
    if isinstance(data, list):
        managers = data
    elif isinstance(data, dict) and "managers" in data:
        managers = data["managers"]
    else:
        managers = [data]

    if not managers:
        ws.cell(row=1, column=1, value="No data to display")
        return

    # Write headers
    ws.cell(row=1, column=1, value="Category / Question").font = header_font
    for col, manager in enumerate(managers, start=2):
        name = manager.get("manager", manager.get("name", f"Manager {col - 1}"))
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font_white
        cell.fill = header_fill

    # Write data rows
    row = 2
    for manager_idx, manager in enumerate(managers):
        answers = manager.get("answers", [])
        for answer in answers:
            if manager_idx == 0:
                ws.cell(row=row, column=1, value=answer.get("question_text", ""))
            cell = ws.cell(
                row=row,
                column=manager_idx + 2,
                value=answer.get("answer", "—"),
            )
            confidence = answer.get("confidence", "")
            if confidence == "LOW":
                cell.fill = PatternFill(
                    start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
                )
            row += 1
        row = 2  # Reset for next manager (filling same rows, different column)

    # Auto-width columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)


def main():
    parser = argparse.ArgumentParser(description="Generate Excel comparison matrix")
    parser.add_argument("input", type=Path, help="Input JSON file")
    parser.add_argument("--output", "-o", type=Path, required=True, help="Output .xlsx file")
    args = parser.parse_args()

    if not args.input.exists():
        print(f"Input file not found: {args.input}", file=sys.stderr)
        sys.exit(1)

    args.output.parent.mkdir(parents=True, exist_ok=True)

    with open(args.input) as f:
        data = json.load(f)

    wb = Workbook()
    build_comparison_matrix(data, wb)
    wb.save(args.output)
    print(f"Excel matrix written to {args.output}")


if __name__ == "__main__":
    main()

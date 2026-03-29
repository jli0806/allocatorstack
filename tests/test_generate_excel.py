"""Tests for scripts/generate-excel.py — Excel comparison matrix generation."""
import importlib.util
import json
import sys
import subprocess
from pathlib import Path

import pytest
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

SCRIPT = Path(__file__).resolve().parent.parent / "scripts" / "generate-excel.py"
sys.path.insert(0, str(SCRIPT.parent))

_spec = importlib.util.spec_from_file_location("generate_excel", SCRIPT)
_mod = importlib.util.module_from_spec(_spec)
_spec.loader.exec_module(_mod)

build_comparison_matrix = _mod.build_comparison_matrix
COLORS = _mod.COLORS


def _run_script(*args: str) -> subprocess.CompletedProcess:
    return subprocess.run(
        [sys.executable, str(SCRIPT), *args],
        capture_output=True,
        text=True,
    )


# ---------------------------------------------------------------------------
# Single manager
# ---------------------------------------------------------------------------

class TestSingleManager:
    def test_creates_xlsx_file(self, single_manager_ddq_json, output_dir):
        """Script should create a valid .xlsx file for single-manager input."""
        out = output_dir / "single.xlsx"
        result = _run_script(str(single_manager_ddq_json), "--output", str(out))

        assert result.returncode == 0
        assert out.exists()
        assert out.stat().st_size > 0

    def test_xlsx_is_valid_and_loadable(self, single_manager_ddq_json, output_dir):
        """The output file should be a valid Excel workbook readable by openpyxl."""
        out = output_dir / "single.xlsx"
        _run_script(str(single_manager_ddq_json), "--output", str(out))

        wb = load_workbook(out)
        assert len(wb.sheetnames) >= 1
        ws = wb.active
        assert ws.title == "Comparison Matrix"
        wb.close()

    def test_single_manager_has_correct_columns(self, single_manager_ddq_json, output_dir):
        """Single-manager output should have 2 columns: questions + 1 manager."""
        out = output_dir / "single.xlsx"
        _run_script(str(single_manager_ddq_json), "--output", str(out))

        wb = load_workbook(out)
        ws = wb.active
        # Header row: col 1 = "Category / Question", col 2 = manager name
        assert ws.cell(row=1, column=1).value == "Category / Question"
        assert ws.cell(row=1, column=2).value == "Granite Peak Capital"
        # Should have data rows
        assert ws.cell(row=2, column=1).value is not None
        wb.close()

    def test_answers_populated(self, single_manager_ddq_json, output_dir):
        """Answer values should appear in the cells."""
        out = output_dir / "single.xlsx"
        _run_script(str(single_manager_ddq_json), "--output", str(out))

        wb = load_workbook(out)
        ws = wb.active
        # Collect all answer values from column 2
        answers = [ws.cell(row=r, column=2).value for r in range(2, 5)]
        assert "$2.5B" in answers
        assert "Long/short equity" in answers
        wb.close()


# ---------------------------------------------------------------------------
# Multi-manager comparison
# ---------------------------------------------------------------------------

class TestMultiManager:
    def test_multi_manager_creates_file(self, multi_manager_ddq_json, output_dir):
        """Multi-manager input should produce a valid Excel file."""
        out = output_dir / "multi.xlsx"
        result = _run_script(str(multi_manager_ddq_json), "--output", str(out))

        assert result.returncode == 0
        assert out.exists()

    def test_multi_manager_has_correct_column_count(self, multi_manager_ddq_json, output_dir):
        """Multi-manager output should have columns for questions + each manager."""
        out = output_dir / "multi.xlsx"
        _run_script(str(multi_manager_ddq_json), "--output", str(out))

        wb = load_workbook(out)
        ws = wb.active
        # col 1 = questions, cols 2-4 = 3 managers
        assert ws.cell(row=1, column=2).value == "Granite Peak Capital"
        assert ws.cell(row=1, column=3).value == "Meridian Value Partners"
        assert ws.cell(row=1, column=4).value == "Osprey Global Advisors"
        wb.close()

    def test_multi_manager_question_column(self, multi_manager_ddq_json, output_dir):
        """The first column should contain the question text."""
        out = output_dir / "multi.xlsx"
        _run_script(str(multi_manager_ddq_json), "--output", str(out))

        wb = load_workbook(out)
        ws = wb.active
        questions = [ws.cell(row=r, column=1).value for r in range(2, 4)]
        assert "What is your AUM?" in questions
        wb.close()

    def test_multi_manager_data_in_correct_columns(self, multi_manager_ddq_json, output_dir):
        """Each manager's data should appear in its own column."""
        out = output_dir / "multi.xlsx"
        _run_script(str(multi_manager_ddq_json), "--output", str(out))

        wb = load_workbook(out)
        ws = wb.active
        # Find the AUM row (row 2 based on the fixture)
        # Manager 1 (col 2): $2.5B, Manager 2 (col 3): $800M, Manager 3 (col 4): $5.1B
        row2_vals = [ws.cell(row=2, column=c).value for c in range(2, 5)]
        assert "$2.5B" in row2_vals
        assert "$800M" in row2_vals
        assert "$5.1B" in row2_vals
        wb.close()


# ---------------------------------------------------------------------------
# Conditional formatting / color coding
# ---------------------------------------------------------------------------

class TestColorCoding:
    def test_low_confidence_highlighted(self, single_manager_ddq_json, output_dir):
        """Cells with LOW confidence should have a yellow-ish fill applied."""
        out = output_dir / "colors.xlsx"
        _run_script(str(single_manager_ddq_json), "--output", str(out))

        wb = load_workbook(out)
        ws = wb.active

        # The 3rd answer ("Key person risk?") has confidence LOW
        # It should be in row 4, column 2
        low_conf_cell = ws.cell(row=4, column=2)
        fill = low_conf_cell.fill
        # The script applies FFF3CD fill for LOW confidence
        assert fill.fgColor is not None or fill.start_color is not None
        # Check the color is the expected warm yellow
        color_rgb = fill.start_color.rgb if fill.start_color else ""
        assert "FFF3CD" in str(color_rgb), f"Expected FFF3CD fill, got {color_rgb}"
        wb.close()

    def test_high_confidence_no_special_fill(self, single_manager_ddq_json, output_dir):
        """Cells with HIGH confidence should NOT have the low-confidence fill."""
        out = output_dir / "colors.xlsx"
        _run_script(str(single_manager_ddq_json), "--output", str(out))

        wb = load_workbook(out)
        ws = wb.active
        # Row 2 has HIGH confidence
        high_conf_cell = ws.cell(row=2, column=2)
        fill = high_conf_cell.fill
        color_rgb = str(fill.start_color.rgb) if fill.start_color else ""
        assert "FFF3CD" not in color_rgb
        wb.close()

    def test_colors_dict_has_expected_keys(self):
        """The COLORS dict should define severity levels used in the codebase."""
        expected = {"CRITICAL", "WARNING", "INFO", "PASS", "FAIL", "BORDERLINE"}
        assert set(COLORS.keys()) == expected

    def test_color_values_are_hex(self):
        """All color values should be valid 6-char hex strings."""
        for key, val in COLORS.items():
            assert len(val) == 6, f"COLORS[{key}] = {val!r} is not 6 chars"
            int(val, 16)  # Should not raise


# ---------------------------------------------------------------------------
# Edge cases
# ---------------------------------------------------------------------------

class TestEdgeCases:
    def test_empty_managers_list(self, tmp_path, output_dir):
        """An empty managers list should produce a workbook with a 'no data' message."""
        data = {"managers": []}
        inp = tmp_path / "empty.json"
        inp.write_text(json.dumps(data))
        out = output_dir / "empty.xlsx"
        result = _run_script(str(inp), "--output", str(out))

        assert result.returncode == 0
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=1, column=1).value == "No data to display"
        wb.close()

    def test_missing_input_file_exits_nonzero(self, output_dir):
        """Script should exit non-zero when input file does not exist."""
        out = output_dir / "missing.xlsx"
        result = _run_script("/tmp/nonexistent-allocator-test.json", "--output", str(out))
        assert result.returncode != 0
        assert "not found" in result.stderr.lower() or "Input file" in result.stderr

    def test_list_format_input(self, tmp_path, output_dir):
        """When input JSON is a list (not a dict), treat each item as a manager."""
        data = [
            {
                "name": "Fund Alpha",
                "answers": [
                    {"question_text": "AUM?", "answer": "$1B", "confidence": "HIGH"},
                ],
            }
        ]
        inp = tmp_path / "list.json"
        inp.write_text(json.dumps(data))
        out = output_dir / "list.xlsx"
        result = _run_script(str(inp), "--output", str(out))

        assert result.returncode == 0
        wb = load_workbook(out)
        ws = wb.active
        assert ws.cell(row=1, column=2).value == "Fund Alpha"
        wb.close()

    def test_output_directory_auto_created(self, tmp_path):
        """The script should create the output directory if it doesn't exist."""
        data = {"manager": "Test", "answers": []}
        inp = tmp_path / "test.json"
        inp.write_text(json.dumps(data))
        out = tmp_path / "nested" / "deep" / "output.xlsx"
        result = _run_script(str(inp), "--output", str(out))

        assert result.returncode == 0
        assert out.exists()
